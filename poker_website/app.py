from flask import Flask, render_template, request
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# Log data to Excel
def log_to_excel(username, buy_in, cash_out, result):
    # Load the existing Excel workbook or create a new one if it doesn't exist
    try:
        workbook = load_workbook('poker_sessions.xlsx')
    except FileNotFoundError:
        from openpyxl import Workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Name", "Buy-In (€)", "Cash-Out (€)", "Net Profit/Loss (€)"])
        workbook.save('poker_sessions.xlsx')

    # Select the active sheet
    sheet = workbook.active

    # Get the current date
    date = datetime.now().strftime("%d-%b")

    # Append the new data to the sheet
    sheet.append([date, username, buy_in, cash_out, result])

    # Save the workbook
    workbook.save('poker_sessions.xlsx')

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/buy_in', methods=['POST'])
def username():
    username = request.form['username']
    return render_template('buy_in.html', username=username)

@app.route('/buy_in/<username>', methods=['POST'])
def buy_in(username):
    buy_in_amount = float(request.form['buy_in'])
    cash_out_amount = float(request.form['cash_out'])
    result = round(cash_out_amount - buy_in_amount, 2)

    if result > 0:
        outcome = f"Congratulations {username}, you won €{result}!"
    elif result < 0:
        outcome = f"Sorry {username}, you lost €{-result}."
    else:
        outcome = f"{username}, you broke even."

    # Log data to the Excel file
    log_to_excel(username, buy_in_amount, cash_out_amount, result)

    return render_template('result.html', username=username, buy_in_amount=buy_in_amount, cash_out_amount=cash_out_amount, result=result, outcome=outcome)

if __name__ == '__main__':
    app.run(debug=True)
